# views.py
from rest_framework import viewsets, mixins,status
from rest_framework.permissions import IsAuthenticated,AllowAny
from .models import LabTest, Profile, Package,LabCategory
from .serializers import LabTestSerializer, ProfileSerializer, PackageSerializer, LabCategorySerializer
from drpathcare.pagination import StandardResultsSetPagination
import pandas as pd
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import filters 
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from openpyxl import load_workbook
from decimal import Decimal




class BaseLabViewSet(viewsets.GenericViewSet):
    """Common filtering logic shared by CRM & Client"""
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]
    ordering_fields = ["id","name", "category__name", "price", "created_at"]

    def get_queryset(self):
        qs = super().get_queryset()
        category = self.request.query_params.get("category")
        if category:
            qs = qs.filter(category_id=category)
        return qs



# CRM = Full CRUD
class LabTestCRMViewSet(BaseLabViewSet, viewsets.ModelViewSet):
    queryset = LabTest.objects.all()
    serializer_class = LabTestSerializer

    @action(detail=False, methods=["post"], url_path="bulk-upload")
    def bulk_upload(self, request):
        """
        Upload Excel file to create multiple LabTests at once
        Expected columns: name, test_code, investigation, sample_type,
        special_instruction, method, reported_on, category_name, price, sample_required
        """
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Excel file is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        errors = []

        for idx, row in df.iterrows():
            try:
                category_name = row.get("category_name")
                category = None
                if category_name:
                    category, _ = LabCategory.objects.get_or_create(name=category_name, entity_type="test")

                lab_test = LabTest.objects.create(
                            name=row.get("name"),
                            test_code=row.get("test_code"),
                            investigation=row.get("investigation"),
                            sample_type=row.get("sample_type"),
                            special_instruction=row.get("special_instruction"),
                            method=row.get("method"),
                            reported_on=row.get("reported_on"),
                            category=category,
                            price=row.get("price") or 0,
                            offer_price=row.get("offer_price") or None,
                            sample_required=row.get("sample_required")
                        )

                created.append(lab_test.id)
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

        return Response({"created": created, "errors": errors}, status=status.HTTP_201_CREATED)



class ProfileCRMViewSet(BaseLabViewSet, viewsets.ModelViewSet):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer


class PackageCRMViewSet(BaseLabViewSet, viewsets.ModelViewSet):
    queryset = Package.objects.all()
    serializer_class = PackageSerializer

class LabCategoryCRMViewSet(BaseLabViewSet, viewsets.ModelViewSet):
    queryset = LabCategory.objects.all()
    serializer_class = LabCategorySerializer

# Client = Read only

class LabCategoryClientViewSet(BaseLabViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin):
    queryset = LabCategory.objects.all().order_by("name")
    serializer_class = LabCategorySerializer
    permission_classes = [AllowAny]
    

class LabTestClientViewSet(BaseLabViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin):
    queryset = LabTest.objects.all()
    serializer_class = LabTestSerializer
    permission_classes = [AllowAny]
    search_fields = ["name","is_featured"]


class ProfileClientViewSet(BaseLabViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer
    permission_classes = [AllowAny]


class PackageClientViewSet(BaseLabViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin):
    queryset = Package.objects.all()
    serializer_class = PackageSerializer
    permission_classes = [AllowAny]
    search_fields = ["name","is_featured"]


class LabCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = LabCategory.objects.all().order_by("name")
    serializer_class = LabCategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name","entity_type","description"]
    ordering_fields = ["name", "entity_type", "description", "created_at"]
    ordering = ["name"]  # default ordering

    pagination_class = StandardResultsSetPagination  # default

    def paginate_queryset(self, queryset):
        page_size = self.request.query_params.get("page_size")
        if page_size is None:
            return None  # disable pagination
        return super().paginate_queryset(queryset)


    def get_queryset(self):
        entity_type = self.request.query_params.get("entity_type")
        qs = super().get_queryset()
        if entity_type:
            qs = qs.filter(entity_type=entity_type)
        return qs


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def global_search(request):
    query = request.GET.get('q', '').strip()
    if not query:
        return Response({"error": "Missing query parameter ?q="}, status=status.HTTP_400_BAD_REQUEST)

    limit = 15
    results = []

    # Search Lab Tests
    lab_tests = (
        LabTest.objects
        .filter(Q(name__icontains=query))
        .values('id', 'name')
        [:5]
    )
    for item in lab_tests:
        results.append({
            "id": item["id"],
            "name": item["name"],
            "type": "LabTest"
        })

    # Search Profiles
    profiles = (
        Profile.objects
        .filter(Q(name__icontains=query))
        .values('id', 'name')
        [:5]
    )
    for item in profiles:
        results.append({
            "id": item["id"],
            "name": item["name"],
            "type": "Profile"
        })

    # Search Packages
    packages = (
        Package.objects
        .filter(Q(name__icontains=query))
        .values('id', 'name')
        [:5]
    )
    for item in packages:
        results.append({
            "id": item["id"],
            "name": item["name"],
            "type": "Package"
        })

    # Combine and trim total results to 15
    results = results[:limit]

    return Response({
        "query": query,
        "count": len(results),
        "results": results
    })


class LabTestBulkUploadAPIView(APIView):
    permission_classes = [IsAuthenticated]  # CRM-only

    REQUIRED_COLUMNS = [
        "name",
        "test_code",
        "sample_type",
        "special_instruction",
        "temperature",
        "method",
        "reported_on",
        "category",
        "price",
        "offer_price",
        "description",
    ]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded"}, status=400)

        if not file.name.endswith(".xlsx"):
            return Response({"error": "Only .xlsx files allowed"}, status=400)

        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response({"error": f"Failed to read Excel file: {e}"}, status=400)

        # --- Read header row ---
        header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        # --- Validate required columns ---
        missing = [col for col in self.REQUIRED_COLUMNS if col not in header]
        if missing:
            return Response(
                {"error": f"Missing required columns: {', '.join(missing)}"},
                status=400
            )

        # Mapping column → index
        col_map = {col: header.index(col) for col in header}

        created_count = 0
        updated_count = 0
        errors = []

        # --- Process rows ---
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

            def get(col):
                return row[col_map[col]].value if col_map[col] < len(row) else None

            # --------------------------
            # Name
            # --------------------------
            name = (get("name") or "").strip()
            if not name:
                errors.append(f"Row {idx}: Missing name")
                continue

            # --------------------------
            # Category
            # --------------------------
            category_name = (get("category") or "").strip()
            if not category_name:
                errors.append(f"Row {idx}: Missing category")
                continue

            category_obj, _ = LabCategory.objects.get_or_create(
                name=category_name,
                entity_type="lab_test"
            )

            # --------------------------
            # Retrieve or create LabTest (SAFE)
            # --------------------------
            try:
                test_obj = LabTest.objects.get(name=name)
                created = False
            except LabTest.DoesNotExist:
                test_obj = LabTest(name=name)
                created = True

            # --------------------------
            # Assign fields (common for update & create)
            # --------------------------
            test_obj.test_code = get("test_code") or ""
            test_obj.sample_type = get("sample_type") or ""
            test_obj.special_instruction = get("special_instruction") or ""
            test_obj.method = get("method") or ""
            test_obj.temperature = get("temperature") or ""
            test_obj.description = get("description") or ""
            test_obj.reported_on = get("reported_on") or ""
            test_obj.category = category_obj

            # --------------------------
            # PRICE (must exist)
            # --------------------------
            price_value = int(get("price"))
            if price_value in [None, ""]:
                errors.append(f"Row {idx}: Missing price")
                continue

            try:
                print(price_value)
                test_obj.price = Decimal(price_value)
            except Exception:
                errors.append(f"Row {idx}: Invalid price")
                break

            # --------------------------
            # OFFER PRICE (optional)
            # --------------------------
            offer_price_val = get("offer_price")
            try:
                test_obj.offer_price = Decimal(offer_price_val) if offer_price_val else None
            except Exception:
                errors.append(f"Row {idx}: Invalid offer_price")
                continue

            # --------------------------
            # Save record
            # --------------------------
            try:
                test_obj.save()
            except Exception as e:
                errors.append(f"Row {idx}: Save failed - {e}")
                continue

            if created:
                created_count += 1
            else:
                updated_count += 1

        # --- Final response ---
        return Response(
            {
                "status": "Bulk upload completed",
                "created": created_count,
                "updated": updated_count,
                "errors": errors,
            },
            status=200,
        )